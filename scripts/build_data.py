from __future__ import annotations

import csv
import json
import math
import re
import unicodedata
from pathlib import Path
from xml.sax.saxutils import escape

import openpyxl


ROOT = Path(__file__).resolve().parents[1]
RAW = ROOT / "data" / "raw"
PUBLIC = ROOT / "public" / "data"
MAP_XLSX = Path(r"C:\workspace\creches\munic_icon_map.xlsx")
CSV_2026 = RAW / "ips_brasil_2026_pe.csv"
CSV_2026_BR = RAW / "ips_brasil_2026_brasil.csv"

DIMENSIONS = [
    "Necessidades Humanas Básicas",
    "Fundamentos do Bem-estar",
    "Oportunidades",
]

COMPONENTS = [
    "Nutrição e Cuidados Médicos Básicos",
    "Água e Saneamento",
    "Moradia",
    "Segurança Pessoal",
    "Acesso ao Conhecimento Básico",
    "Acesso à Informação e Comunicação",
    "Saúde e Bem-estar",
    "Qualidade do Meio Ambiente",
    "Direitos Individuais",
    "Liberdades Individuais e de Escolha",
    "Inclusão Social",
    "Acesso à Educação Superior",
]

LOWER_IS_BETTER = {
    "Hospitaliza\u00e7\u00f5es por Condi\u00e7\u00f5es Sens\u00edveis \u00e0 Aten\u00e7\u00e3o Prim\u00e1ria",
    "Mortalidade Ajustada por Condi\u00e7\u00f5es Sens\u00edveis \u00e0 Aten\u00e7\u00e3o Prim\u00e1ria",
    "Mortalidade Infantil at\u00e9 5 anos",
    "Subnutri\u00e7\u00e3o",
    "\u00cdndice de Perdas de \u00c1gua na Distribui\u00e7\u00e3o",
    "Assassinatos de Jovens",
    "Assassinatos de Mulheres",
    "Homic\u00eddios",
    "Mortes por Acidentes de Transporte",
    "Abandono no Ensino Fundamental",
    "Abandono no Ensino M\u00e9dio",
    "Distor\u00e7\u00e3o Idade-S\u00e9rie no Ensino M\u00e9dio",
    "Evas\u00e3o no Ensino M\u00e9dio",
    "Reprova\u00e7\u00e3o Escolar no Ensino M\u00e9dio",
    "Consumo de ultraprocessados",
    "Mortalidade entre 15 e 50 anos",
    "Mortalidades por Doen\u00e7as Cr\u00f4nicas N\u00e3o Transmiss\u00edveis",
    "Obesidade",
    "Suic\u00eddios",
    "Emiss\u00f5es de CO\u2082e por Habitante",
    "Focos de Calor",
    "\u00cdndice de Vulnerabilidade Clim\u00e1tica dos Munic\u00edpios (IVCM)",
    "Supress\u00e3o da Vegeta\u00e7\u00e3o Prim\u00e1ria e Secund\u00e1ria",
    "Taxa de Congestionamento L\u00edquida de Processos",
    "Gravidez na Adolesc\u00eancia (<19)",
    "\u00cdndice de Vulnerabilidade das Fam\u00edlias do Cad\u00danico (IVCAD)",
    "Fam\u00edlias em Situa\u00e7\u00e3o de Rua",
    "Viol\u00eancia Contra Ind\u00edgenas",
    "Viol\u00eancia Contra Mulheres",
    "Viol\u00eancia Contra Negros",
}
ALIASES = {
    "ÁREA km²": "Área (km²)",
    "POPULAÇÃO 2025": "População",
    "População 2022": "População",
    "PIB PER CAPITA": "PIB per capita",
    "PIB per capita 2021": "PIB per capita",
    "Cobertura Vacinal (Poliomielite)": "Cobertura Vacinal (poliomielite)",
    "Mortalidade Infantil até 5 Anos": "Mortalidade Infantil até 5 anos",
    "Abastecimento de Água via Rede de Distribuição": "Abastecimento de Água Via Rede de Distribuição",
    "Domicílios com Piso Adequado": "Domicílios com Pisos Adequados",
    "Mortes por Acidente de Transporte": "Mortes por Acidentes de Transporte",
    "Densidade de Telefonia Móvel": "Densidade Telefonia Móvel",
    "Consumo de Alimentos Ultraprocessados": "Consumo de ultraprocessados",
    "Mortalidade por Doenças Crônicas Não Transmissíveis": "Mortalidades por Doenças Crônicas Não Transmissíveis",
    "Taxa de Congestionamento Líquido de Processos": "Taxa de Congestionamento Líquida de Processos",
    "Gravidez na Adolescência (<19 anos)": "Gravidez na Adolescência (<19)",
    "Índice de Vulnerabilidade das Famílias do Cadastro Único (IVCAD)": "Índice de Vulnerabilidade das Famílias do CadÚnico (IVCAD)",
    "Violência contra Indígenas": "Violência Contra Indígenas",
    "Violência contra Mulheres": "Violência Contra Mulheres",
    "Violência contra Negros": "Violência Contra Negros",
    "Nota Mediana do Enem": "Nota Mediana no Enem",
}

NAME_ALIASES = {
    "iguaracy": "iguaraci",
    "lagoadeitaenga": "lagoadoitaenga",
    "belemdosaofrancisco": "belemdesaofrancisco",
}

UF_TO_REGION = {
    "AC": "Norte",
    "AL": "Nordeste",
    "AM": "Norte",
    "AP": "Norte",
    "BA": "Nordeste",
    "CE": "Nordeste",
    "DF": "Centro-Oeste",
    "ES": "Sudeste",
    "GO": "Centro-Oeste",
    "MA": "Nordeste",
    "MG": "Sudeste",
    "MS": "Centro-Oeste",
    "MT": "Centro-Oeste",
    "PA": "Norte",
    "PB": "Nordeste",
    "PE": "Nordeste",
    "PI": "Nordeste",
    "PR": "Sul",
    "RJ": "Sudeste",
    "RN": "Nordeste",
    "RO": "Norte",
    "RR": "Norte",
    "RS": "Sul",
    "SC": "Sul",
    "SE": "Nordeste",
    "SP": "Sudeste",
    "TO": "Norte",
}

CAPITALS = {
    "1200401": ("Rio Branco", "AC"),
    "2704302": ("Maceió", "AL"),
    "1302603": ("Manaus", "AM"),
    "1600303": ("Macapá", "AP"),
    "2927408": ("Salvador", "BA"),
    "2304400": ("Fortaleza", "CE"),
    "5300108": ("Brasília", "DF"),
    "3205309": ("Vitória", "ES"),
    "5208707": ("Goiânia", "GO"),
    "2111300": ("São Luís", "MA"),
    "3106200": ("Belo Horizonte", "MG"),
    "5002704": ("Campo Grande", "MS"),
    "5103403": ("Cuiabá", "MT"),
    "1501402": ("Belém", "PA"),
    "2507507": ("João Pessoa", "PB"),
    "2611606": ("Recife", "PE"),
    "2211001": ("Teresina", "PI"),
    "4106902": ("Curitiba", "PR"),
    "3304557": ("Rio De Janeiro", "RJ"),
    "2408102": ("Natal", "RN"),
    "1100205": ("Porto Velho", "RO"),
    "1400100": ("Boa Vista", "RR"),
    "4314902": ("Porto Alegre", "RS"),
    "4205407": ("Florianópolis", "SC"),
    "2800308": ("Aracaju", "SE"),
    "3550308": ("São Paulo", "SP"),
    "1721000": ("Palmas", "TO"),
}
CAPITAL_CODES = set(CAPITALS)


def slug(value: str) -> str:
    value = unicodedata.normalize("NFKD", value or "")
    value = "".join(ch for ch in value if not unicodedata.combining(ch))
    value = re.sub(r"\s*\([A-Z]{2}\)$", "", value)
    return re.sub(r"[^a-z0-9]+", "", value.lower())


def municipality_key(value: str) -> str:
    key = slug(value)
    return NAME_ALIASES.get(key, key)


CAPITAL_BY_NAME_UF = {(municipality_key(name), uf): code for code, (name, uf) in CAPITALS.items()}


def clean_header(value: str) -> str:
    value = (value or "").strip()
    value = ALIASES.get(value, value)
    return value.strip()


def num(value):
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)):
        if isinstance(value, float) and math.isnan(value):
            return None
        return round(float(value), 6)
    text = str(value).strip()
    if not text:
        return None
    text = text.replace(".", "").replace(",", ".") if re.search(r",\d+$", text) else text
    try:
        return round(float(text), 6)
    except ValueError:
        return None


def load_regions():
    wb = openpyxl.load_workbook(MAP_XLSX, read_only=True, data_only=True)
    ws = wb["munic_icon_map"]
    headers = [str(c) for c in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
    idx = {name: headers.index(name) for name in headers}
    by_name = {}
    by_code = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        item = {
            "code": str(row[idx["cd_mun"]]),
            "name": str(row[idx["mun_sem_ac"]] or row[idx["mun_com_ac"]]),
            "region": title_region(str(row[idx["rd_nome"]])),
            "wkt": row[idx["geom_prec_reduz"]] or row[idx["geom"]],
        }
        by_name[municipality_key(item["name"])] = item
        by_code[item["code"]] = item
    return by_name, by_code


def title_region(value: str) -> str:
    fixed = {
        "REGIAO": "Região",
        "SERTAO": "Sertão",
        "SAO": "São",
        "MOXOTO": "Moxotó",
        "PAJEU": "Pajeú",
    }
    parts = []
    for word in value.split():
        parts.append(fixed.get(word, word.capitalize()))
    return " ".join(parts)


def read_csv(path: Path, year: int, by_name, by_code):
    with path.open("r", encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f)
        rows = []
        for source in reader:
            source = {clean_header(k): v for k, v in source.items()}
            code = source.get("Código IBGE")
            name = re.sub(r"\s*\([A-Z]{2}\)$", "", source["Município"]).strip()
            meta = by_code.get(str(code)) if code else by_name.get(municipality_key(name))
            if not meta:
                raise RuntimeError(f"Município não encontrado no mapa: {name}")
            measures = {}
            for key, value in source.items():
                clean = clean_header(key)
                if clean in {"Código IBGE", "Município", "UF"}:
                    continue
                measures[clean] = num(value)
            rows.append(
                {
                    "year": year,
                    "code": meta["code"],
                    "municipality": name.title(),
                    "uf": "PE",
                    "region": meta["region"],
                    "values": measures,
                    "ranks": {},
                }
            )
        return rows


def read_xlsx_rows(path: Path, year: int):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    ws.reset_dimensions()
    iterator = ws.iter_rows(values_only=True)
    headers = [clean_header(str(cell)) for cell in next(iterator)]
    rows = []
    for row in iterator:
        values = list(row)
        code = str(values[0])
        name = re.sub(r"\s*\([A-Z]{2}\)$", "", str(values[1] or "")).strip()
        uf = str(values[2] or "").strip()
        measures = {}
        for key, value in zip(headers[3:], values[3:]):
            measures[clean_header(key)] = num(value)
        rows.append({"year": year, "code": code, "municipality": name.title(), "uf": uf, "values": measures})
    return rows


def read_national_csv_rows(path: Path, year: int):
    with path.open("r", encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f)
        rows = []
        for source in reader:
            source = {clean_header(k): v for k, v in source.items()}
            name = re.sub(r"\s*\([A-Z]{2}\)$", "", source["Município"]).strip()
            uf = source["UF"].strip()
            code = CAPITAL_BY_NAME_UF.get((municipality_key(name), uf), "")
            measures = {}
            for key, value in source.items():
                clean = clean_header(key)
                if clean in {"Município", "UF"}:
                    continue
                measures[clean] = num(value)
            rows.append({"year": year, "code": code, "municipality": name.title(), "uf": uf, "values": measures})
        return rows


def national_row_key(row: dict) -> str:
    if row.get("code"):
        return row["code"]
    return f"{row.get('uf', '')}:{municipality_key(row.get('municipality', ''))}"


def rank_map(rows, indicators, key_fn=lambda row: row["code"]):
    ranks = {}
    for indicator in indicators:
        valid = [row for row in rows if row["values"].get(indicator) is not None]
        valid.sort(key=lambda row: row["values"][indicator], reverse=indicator not in LOWER_IS_BETTER)
        for index, row in enumerate(valid, start=1):
            ranks.setdefault(key_fn(row), {})[indicator] = {"position": index, "total": len(valid)}
    return ranks


def load_national_by_year():
    return {
        2024: read_xlsx_rows(ROOT / "2024_br.xlsx", 2024),
        2025: read_xlsx_rows(ROOT / "2025_br.xlsx", 2025),
        2026: read_national_csv_rows(CSV_2026_BR, 2026),
    }


def attach_ranks(records, national_by_year):
    ranked_indicators = sorted({key for row in records for key in row["values"]})
    by_year = {}
    for row in records:
        by_year.setdefault(row["year"], []).append(row)

    national_ranks = {year: rank_map(rows, ranked_indicators, national_row_key) for year, rows in national_by_year.items()}
    pe_ranks = {year: rank_map(rows, ranked_indicators) for year, rows in by_year.items()}

    for row in records:
        year = row["year"]
        for indicator in ranked_indicators:
            national_key = row["code"] if year != 2026 or row["code"] in CAPITAL_CODES else f"PE:{municipality_key(row['municipality'])}"
            br = national_ranks.get(year, {}).get(national_key, {}).get(indicator)
            pe = pe_ranks.get(year, {}).get(row["code"], {}).get(indicator)
            row["ranks"][indicator] = {"br": br, "pe": pe}


def build_capital_records(national_by_year, indicators):
    records = []
    for year, rows in national_by_year.items():
        capital_rows = []
        for row in rows:
            code = row.get("code") or CAPITAL_BY_NAME_UF.get((municipality_key(row.get("municipality", "")), row.get("uf", "")))
            if code not in CAPITAL_CODES:
                continue
            capital_name, uf = CAPITALS[code]
            capital_rows.append(
                {
                    "year": year,
                    "code": code,
                    "municipality": capital_name,
                    "uf": uf,
                    "region": UF_TO_REGION[uf],
                    "values": row["values"],
                    "ranks": {},
                }
            )

        br_ranks = rank_map(rows, indicators, national_row_key)
        capital_ranks = rank_map(capital_rows, indicators)
        regional_ranks = {}
        for region in sorted(UF_TO_REGION.values()):
            region_rows = [row for row in capital_rows if row["region"] == region]
            regional_ranks[region] = rank_map(region_rows, indicators)

        for row in capital_rows:
            source_key = row["code"]
            for indicator in indicators:
                row["ranks"][indicator] = {
                    "br": br_ranks.get(source_key, {}).get(indicator),
                    "capital": capital_ranks.get(row["code"], {}).get(indicator),
                    "capitalRegion": regional_ranks.get(row["region"], {}).get(row["code"], {}).get(indicator),
                }
        records.extend(sorted(capital_rows, key=lambda item: item["municipality"]))
    return records


def parse_wkt_polygons(wkt: str):
    text = wkt.strip()
    if text.startswith("MULTIPOLYGON"):
        body = text[len("MULTIPOLYGON") :].strip()[1:-1]
        poly_texts = split_top_level(body)
    elif text.startswith("POLYGON"):
        poly_texts = [text[len("POLYGON") :].strip()]
    else:
        return []
    polygons = []
    for poly in poly_texts:
        poly = poly.strip()
        if poly.startswith("(") and poly.endswith(")"):
            poly = poly[1:-1]
        rings = []
        for ring in split_top_level(poly):
            ring = ring.strip()
            if ring.startswith("(") and ring.endswith(")"):
                ring = ring[1:-1]
            points = []
            for pair in ring.split(","):
                parts = pair.strip().split()
                if len(parts) >= 2:
                    points.append((float(parts[0]), float(parts[1])))
            if points:
                rings.append(points)
        if rings:
            polygons.append(rings)
    return polygons


def split_top_level(text: str):
    parts, depth, start = [], 0, 0
    for index, char in enumerate(text):
        if char == "(":
            depth += 1
        elif char == ")":
            depth -= 1
        elif char == "," and depth == 0:
            parts.append(text[start:index])
            start = index + 1
    parts.append(text[start:])
    return parts


def build_svg_paths(by_code):
    polygons_by_code = {code: parse_wkt_polygons(meta["wkt"]) for code, meta in by_code.items()}
    xs, ys = [], []
    for polygons in polygons_by_code.values():
        for rings in polygons:
            for ring in rings:
                for x, y in ring:
                    xs.append(x)
                    ys.append(y)
    min_x, max_x, min_y, max_y = min(xs), max(xs), min(ys), max(ys)
    width = 1000
    height = (max_y - min_y) / (max_x - min_x) * width

    def project(point):
        x, y = point
        px = (x - min_x) / (max_x - min_x) * width
        py = (max_y - y) / (max_y - min_y) * height
        return px, py

    paths = {}
    for code, polygons in polygons_by_code.items():
        chunks = []
        for rings in polygons:
            for ring in rings:
                coords = [project(point) for point in ring]
                if not coords:
                    continue
                d = [f"M {coords[0][0]:.2f} {coords[0][1]:.2f}"]
                d.extend(f"L {x:.2f} {y:.2f}" for x, y in coords[1:])
                d.append("Z")
                chunks.append(" ".join(d))
        paths[code] = " ".join(chunks)
    return {"viewBox": f"0 0 {width:.2f} {height:.2f}", "paths": paths}


def main():
    PUBLIC.mkdir(parents=True, exist_ok=True)
    by_name, by_code = load_regions()
    records = []
    records.extend(read_csv(RAW / "ips_brasil_2024_pe.csv", 2024, by_name, by_code))
    records.extend(read_csv(RAW / "ips_brasil_2025_pe.csv", 2025, by_name, by_code))
    records.extend(read_csv(CSV_2026, 2026, by_name, by_code))
    national_by_year = load_national_by_year()
    attach_ranks(records, national_by_year)

    indicators = sorted({key for row in records for key in row["values"]})
    capital_records = build_capital_records(national_by_year, indicators)
    payload = {
        "records": records,
        "capitalRecords": capital_records,
        "capitalRegions": sorted({row["region"] for row in capital_records}),
        "regions": sorted({row["region"] for row in records}),
        "indicators": indicators,
        "indicatorPolarity": {indicator: ("lower" if indicator in LOWER_IS_BETTER else "higher") for indicator in indicators},
        "dimensions": DIMENSIONS,
        "components": COMPONENTS,
        "map": build_svg_paths(by_code),
        "sourceNote": "2024 e 2025 foram obtidos da instância histórica do IPS Brasil; 2026 foi carregado do CSV local enviado pelo usuário.",
    }
    with (PUBLIC / "dashboard.json").open("w", encoding="utf-8") as f:
        json.dump(payload, f, ensure_ascii=False, separators=(",", ":"))

    print(f"Gerado {PUBLIC / 'dashboard.json'} com {len(records)} registros e {len(capital_records)} registros de capitais.")


if __name__ == "__main__":
    main()
